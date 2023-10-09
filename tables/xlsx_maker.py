from openpyxl import Workbook
from django.http import HttpResponse
from datetime import datetime


def xlsx_maker(table_queryset, columns, device):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-{device}.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'), device=device,
    )
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = f'{device}'

    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for data in table_queryset:
        row_num += 1

        row = [data.get_field(column) for column in columns]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response


